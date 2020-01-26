# 现在我们尝试一下创建一个Excel
from openpyxl import Workbook
# 得到一个对象
wb = Workbook()
# 创建一个Sheet,通常不需要，因为我们有现成的可以使用
# wb.create_sheet('Sheet1')
# 现在 在里面写入数据
ws = wb.active

ws['A1'] = 111

wb.save('1.xlsx')lsls
