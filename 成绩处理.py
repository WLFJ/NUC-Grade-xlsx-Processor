'''
2020@Wlfj.fun

'''
from openpyxl import Workbook, load_workbook

# 源数据路径
file_path = "2018成绩.xlsx"
# 输出文件名称
out_file_name = "out.xlsx"

# 只需要输入绩点不是1的科目
grade_points = {

    "毛泽东思想和中国特色社会主义理论体系概论" : 6,
    "大学英语A（3）" : 4,
    "概率论与数理统计C" : 3,
    "面向对象程序设计" : 3,
    "面向对象程序设计实验" : 1.5,
    "软件工程导论" : 2.5,
    "计算机硬件基础" : 4.5,
    "面向对象程序设计实训" : 2,

}

class Student:
    def __init__(self, id, name):
        self.id = id
        self.name = name
        self.course = {}

    def get_points(self):
        # 参考资料：http://ss.nuc.edu.cn/info/1027/2859.htm

        sum_of_points = 0.0
        total_points = 0

        for course_name in self.course:
            grade = self.course[course_name]
            if grade < 50:
                grade = 50
            sum_of_points += round(grade / 10 - 5, 2) * grade_points.get(course_name, 1)
            total_points += grade_points.get(course_name, 1)
        
        ans = sum_of_points / total_points

        return ans

# 以下部分不需要改动
wb = load_workbook(file_path)

ws = wb.active

database = {}
course_list = set()
for row in ws.rows:
    if row[0].value == '课程号':
        continue
    
    s_id = int(row[1].value)
    s_name = row[4].value
    s_course_name = row[-2].value
    s_course_score = 0

    if row[2].value != None:
        s_course_score = int(row[2].value)

    course_list.add(s_course_name)
    
    if s_id not in database:
        database[s_id] = Student(id = s_id, name = s_name)
    
    stu = database[s_id]
    stu.course[s_course_name] = s_course_score

out_wb = Workbook()

out_ws = out_wb.active

out_ws.cell(row = 1, column = 1, value = "学号")
out_ws.cell(row = 1, column = 2, value = "姓名")

course_pos = {}

course_list = list(course_list)

for i in range(len(course_list)):
    out_ws.cell(row = 1, column = i + 3, value = course_list[i])
    course_pos[course_list[i]] = i + 3

cur_row = 2
for stu_id in database:
    stu = database[stu_id]
    stu_name = stu.name
    out_ws.cell(row = cur_row, column = 1, value = stu_id)
    out_ws.cell(row = cur_row, column = 2, value = stu_name)
    for course_name in stu.course:
        out_ws.cell(row = cur_row, column = course_pos[course_name], value = stu.course[course_name])
    out_ws.cell(row = cur_row, column = len(course_pos) + 3, value = stu.get_points())
    cur_row += 1

out_wb.save(out_file_name)
