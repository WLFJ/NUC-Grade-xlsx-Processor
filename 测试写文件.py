# 现在我们要遍历一个表格
from openpyxl import Workbook, load_workbook
wb = load_workbook('1.xlsx')
ws = wb.active
print(ws['A1'].value)